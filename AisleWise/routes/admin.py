import inspect
from functools import wraps
from datetime import datetime
import csv
import io
import openpyxl
import re
import os
from fastapi import APIRouter, Request, UploadFile, File
from database.db import get_db, query_db
from routes import RedirectException, flash

router = APIRouter()


def admin_login_required(func):
    @wraps(func)
    async def wrapper(*args, **kwargs):
        request = kwargs.get("request") or (args[0] if args else None)
        if not request or not request.session.get("admin_id"):
            raise RedirectException(url="/admin/login")
        if inspect.iscoroutinefunction(func):
            return await func(*args, **kwargs)
        return func(*args, **kwargs)
    return wrapper


@router.get("/", name="admin.index")
def index(request: Request):
    if request.session.get("admin_id"):
        raise RedirectException(url=str(request.url_for("admin.dashboard")))
    raise RedirectException(url=str(request.url_for("admin.login")))


@router.api_route("/login", methods=["GET", "POST"], name="admin.login")
async def login(request: Request):
    error = None
    if request.method == "POST":
        form_data = await request.form()
        store_id = form_data.get("store_id", "").strip()
        password = form_data.get("password", "").strip()
        
        # Verify against updated 'admins' table
        admin = query_db(
            "SELECT * FROM admins WHERE store_id = ? AND password = ?",
            (store_id, password),
            one=True
        )
        if admin is None:
            error = "Invalid store ID or password."
        else:
            request.session.clear()
            request.session["admin_id"] = admin["id"]
            request.session["store_id"] = admin["store_id"]
            raise RedirectException(url=str(request.url_for("admin.dashboard")))
            
    templates = request.app.state.templates
    return templates.TemplateResponse(request=request, name="admin/login.html", context={"error": error})


@router.get("/dashboard", name="admin.dashboard")
@admin_login_required
def dashboard(request: Request):
    total_products = query_db("SELECT COUNT(1) AS count FROM products", one=True)["count"]
    out_of_stock = query_db("SELECT COUNT(1) AS count FROM products WHERE stock = 0", one=True)["count"]
    active_workers = query_db("SELECT COUNT(1) AS count FROM workers WHERE status = 'Active'", one=True)["count"]
    pending_tasks = query_db("SELECT COUNT(1) AS count FROM tasks WHERE status = 'Pending'", one=True)["count"]
    
    low_stock = query_db(
        "SELECT * FROM products WHERE stock < 5 ORDER BY stock ASC"
    )
    
    templates = request.app.state.templates
    return templates.TemplateResponse(
        request=request,
        name="admin/dashboard.html",
        context={
            "total_products": total_products,
            "out_of_stock": out_of_stock,
            "active_workers": active_workers,
            "pending_tasks": pending_tasks,
            "low_stock": low_stock
        }
    )


# ==========================================
# PRODUCT CRUD
# ==========================================

@router.api_route("/inventory", methods=["GET", "POST"], name="admin.inventory")
@admin_login_required
async def inventory(request: Request):
    if request.method == "POST":
        form_data = await request.form()
        product_id = form_data.get("product_id")
        stock = form_data.get("stock")
        if product_id and stock is not None:
            try:
                stock_value = int(stock)
                if stock_value < 0:
                    flash(request, "Stock level cannot be negative.")
                else:
                    db = get_db()
                    db.execute(
                        "UPDATE products SET stock = ? WHERE id = ?",
                        (stock_value, product_id)
                    )
                    db.commit()
                    flash(request, "Stock updated successfully.")
            except ValueError:
                flash(request, "Stock must be a valid number.")
        raise RedirectException(url=str(request.url_for("admin.inventory")))

    products = query_db("SELECT * FROM products ORDER BY id")
    templates = request.app.state.templates
    return templates.TemplateResponse(request=request, name="admin/inventory.html", context={"products": products})


@router.api_route("/add-product", methods=["GET", "POST"], name="admin.add_product")
@admin_login_required
async def add_product(request: Request):
    error = None
    if request.method == "POST":
        form_data = await request.form()
        name = form_data.get("name", "").strip()
        category = form_data.get("category", "").strip()
        price = form_data.get("price", "").strip()
        stock = form_data.get("stock", "").strip()
        aisle = form_data.get("aisle", "").strip()
        description = form_data.get("description", "").strip()

        if not (name and category and price and stock and aisle):
            error = "Name, Category, Price, Stock, and Aisle fields are required."
        else:
            try:
                price_value = float(price)
                stock_value = int(stock)
                if price_value < 0 or stock_value < 0:
                    error = "Price and stock levels cannot be negative."
                else:
                    db = get_db()
                    db.execute(
                        "INSERT INTO products (name, category, price, stock, aisle, description) VALUES (?, ?, ?, ?, ?, ?)",
                        (name, category, price_value, stock_value, aisle, description)
                    )
                    db.commit()
                    flash(request, "Product added to inventory successfully.")
                    raise RedirectException(url=str(request.url_for("admin.inventory")))
            except ValueError:
                error = "Price must be a float and stock must be an integer."

    templates = request.app.state.templates
    return templates.TemplateResponse(request=request, name="admin/add_product.html", context={"error": error})


@router.api_route("/product/edit/{product_id:int}", methods=["GET", "POST"], name="admin.edit_product")
@admin_login_required
async def edit_product(request: Request, product_id: int):
    product = query_db("SELECT * FROM products WHERE id = ?", (product_id,), one=True)
    if not product:
        flash(request, "Product not found.")
        raise RedirectException(url=str(request.url_for("admin.inventory")))
        
    error = None
    if request.method == "POST":
        form_data = await request.form()
        name = form_data.get("name", "").strip()
        category = form_data.get("category", "").strip()
        price = form_data.get("price", "").strip()
        stock = form_data.get("stock", "").strip()
        aisle = form_data.get("aisle", "").strip()
        description = form_data.get("description", "").strip()

        if not (name and category and price and stock and aisle):
            error = "All fields except description are required."
        else:
            try:
                price_val = float(price)
                stock_val = int(stock)
                if price_val < 0 or stock_val < 0:
                    error = "Price and stock levels cannot be negative."
                else:
                    db = get_db()
                    db.execute(
                        """
                        UPDATE products 
                        SET name = ?, category = ?, price = ?, stock = ?, aisle = ?, description = ? 
                        WHERE id = ?
                        """,
                        (name, category, price_val, stock_val, aisle, description, product_id)
                    )
                    db.commit()
                    flash(request, f"Product {name} updated successfully.")
                    raise RedirectException(url=str(request.url_for("admin.inventory")))
            except ValueError:
                error = "Price must be a number and stock must be an integer."
                
    templates = request.app.state.templates
    return templates.TemplateResponse(request=request, name="admin/edit_product.html", context={"product": product, "error": error})


@router.post("/product/delete/{product_id:int}", name="admin.delete_product")
@admin_login_required
def delete_product(request: Request, product_id: int):
    db = get_db()
    product = query_db("SELECT * FROM products WHERE id = ?", (product_id,), one=True)
    if product:
        db.execute("DELETE FROM products WHERE id = ?", (product_id,))
        # Also clean up any pending tasks associated with this product
        db.execute("DELETE FROM tasks WHERE product_id = ?", (product_id,))
        db.commit()
        flash(request, f"Product {product['name']} deleted from inventory.")
    else:
        flash(request, "Product not found.")
    raise RedirectException(url=str(request.url_for("admin.inventory")))


# ==========================================
# WORKER MANAGEMENT
# ==========================================

@router.get("/workers", name="admin.workers")
@admin_login_required
def workers(request: Request):
    workers = query_db("SELECT * FROM workers ORDER BY id")
    templates = request.app.state.templates
    return templates.TemplateResponse(request=request, name="admin/workers.html", context={"workers": workers})


@router.api_route("/workers/add", methods=["GET", "POST"], name="admin.add_worker")
@admin_login_required
async def add_worker(request: Request):
    error = None
    if request.method == "POST":
        form_data = await request.form()
        name = form_data.get("name", "").strip()
        email = form_data.get("email", "").strip().lower()
        status = form_data.get("status", "Inactive")
        
        if not (name and email):
            error = "Name and email are required."
        elif not email.endswith("@gmail.com"):
            error = "Worker email must be an approved Gmail account (ending in @gmail.com)."
        else:
            try:
                db = get_db()
                db.execute(
                    "INSERT INTO workers (name, email, status) VALUES (?, ?, ?)",
                    (name, email, status)
                )
                db.commit()
                flash(request, f"Worker account for {name} registered.")
                raise RedirectException(url=str(request.url_for("admin.workers")))
            except Exception as e:
                error = "Unable to add worker (email may already exist)."
    templates = request.app.state.templates
    return templates.TemplateResponse(request=request, name="admin/add_worker.html", context={"error": error})


@router.post("/workers/toggle/{worker_id:int}", name="admin.toggle_worker")
@admin_login_required
def toggle_worker(request: Request, worker_id: int):
    db = get_db()
    row = query_db("SELECT * FROM workers WHERE id = ?", (worker_id,), one=True)
    if row:
        new_status = "Active" if row["status"] != "Active" else "Inactive"
        db.execute("UPDATE workers SET status = ? WHERE id = ?", (new_status, worker_id))
        db.commit()
        flash(request, f"Worker status toggled to {new_status}.")
    raise RedirectException(url=str(request.url_for("admin.workers")))


# ==========================================
# WORKER TASK ASSIGNMENT
# ==========================================

@router.get("/tasks", name="admin.tasks")
@admin_login_required
def tasks(request: Request):
    assigned_tasks = query_db(
        """
        SELECT t.id AS task_id, t.description AS task_desc, t.status AS task_status, t.timestamp AS task_time,
               w.name AS worker_name, p.name AS product_name, p.aisle AS product_aisle
        FROM tasks t
        JOIN workers w ON t.worker_id = w.id
        JOIN products p ON t.product_id = p.id
        ORDER BY t.timestamp DESC
        """
    )
    templates = request.app.state.templates
    return templates.TemplateResponse(request=request, name="admin/tasks.html", context={"tasks": assigned_tasks})


@router.api_route("/tasks/assign", methods=["GET", "POST"], name="admin.assign_task")
@admin_login_required
async def assign_task(request: Request):
    error = None
    db = get_db()
    
    if request.method == "POST":
        form_data = await request.form()
        worker_id = form_data.get("worker_id")
        product_id = form_data.get("product_id")
        description = form_data.get("description", "").strip()
        
        if not (worker_id and product_id and description):
            error = "All fields are required to assign a task."
        else:
            try:
                timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                db.execute(
                    "INSERT INTO tasks (worker_id, product_id, description, status, timestamp) VALUES (?, ?, ?, ?, ?)",
                    (worker_id, product_id, description, 'Pending', timestamp)
                )
                db.commit()
                flash(request, "Task successfully assigned to worker.")
                raise RedirectException(url=str(request.url_for("admin.tasks")))
            except Exception as e:
                error = f"Error assigning task: {e}"
                
    # Fetch data for dropdowns
    active_workers = query_db("SELECT id, name FROM workers WHERE status = 'Active'")
    products = query_db("SELECT id, name, stock, aisle FROM products ORDER BY name ASC")
    
    templates = request.app.state.templates
    return templates.TemplateResponse(
        request=request,
        name="admin/assign_task.html",
        context={
            "active_workers": active_workers,
            "products": products,
            "error": error
        }
    )


# ==========================================
# QUERY & INVENTORY MONITORING ANALYTICS
# ==========================================

@router.get("/analytics", name="admin.analytics")
@admin_login_required
def analytics(request: Request):
    customer_queries = query_db("SELECT query, response, timestamp FROM customer_queries ORDER BY timestamp DESC LIMIT 50")
    activity_logs = query_db(
        """
        SELECT l.worker_email, l.action, l.timestamp, p.name AS product_name 
        FROM activity_logs l
        LEFT JOIN products p ON l.product_id = p.id
        ORDER BY l.timestamp DESC LIMIT 50
        """
    )
    
    templates = request.app.state.templates
    return templates.TemplateResponse(
        request=request,
        name="admin/analytics.html",
        context={
            "customer_queries": customer_queries,
            "activity_logs": activity_logs
        }
    )


@router.get("/logout", name="admin.logout")
@admin_login_required
def logout(request: Request):
    request.session.clear()
    raise RedirectException(url=str(request.url_for("admin.login")))


# ==========================================
# DATA MANAGEMENT & WORKER PERMISSIONS
# ==========================================

@router.get("/data-management", name="admin.data_management")
@admin_login_required
def data_management(request: Request):
    # Fetch customer display settings
    display_settings = query_db("SELECT * FROM display_settings ORDER BY id")
    
    # Fetch uploaded files history
    uploaded_files = query_db("SELECT * FROM uploaded_files ORDER BY upload_date DESC")
    
    # Fetch custom columns
    db = get_db()
    cursor = db.execute("PRAGMA table_info(products)")
    columns = cursor.fetchall()
    default_fields = {"id", "name", "category", "price", "stock", "aisle", "description"}
    custom_columns = []
    for col in columns:
        col_name = col["name"]
        if col_name not in default_fields:
            custom_columns.append({
                "name": col_name,
                "type": col["type"]
            })
            
    import_summary = request.session.pop("import_summary", None)
    
    templates = request.app.state.templates
    return templates.TemplateResponse(
        request=request,
        name="admin/data_management.html",
        context={
            "display_settings": display_settings,
            "uploaded_files": uploaded_files,
            "custom_columns": custom_columns,
            "import_summary": import_summary
        }
    )


@router.post("/upload-file", name="admin.upload_file")
@admin_login_required
async def upload_file(request: Request, file: UploadFile = File(...)):
    filename = file.filename
    if not (filename.endswith(".csv") or filename.endswith(".xlsx")):
        flash(request, "Invalid file format. Only .csv and .xlsx files are accepted.")
        raise RedirectException(url=str(request.url_for("admin.data_management")))
        
    try:
        contents = await file.read()
        rows = []
        if filename.endswith(".xlsx"):
            wb = openpyxl.load_workbook(io.BytesIO(contents), data_only=True)
            sheet = wb.active
            for row in sheet.iter_rows(values_only=True):
                if any(cell is not None for cell in row):
                    rows.append([str(cell) if cell is not None else "" for cell in row])
        else:
            decoded = contents.decode("utf-8", errors="ignore")
            csv_reader = csv.reader(io.StringIO(decoded))
            rows = list(csv_reader)
            
        if not rows:
            flash(request, "The uploaded file is empty.")
            raise RedirectException(url=str(request.url_for("admin.data_management")))
            
        headers = [h.strip() for h in rows[0] if h is not None]
        
        temp_dir = "database"
        os.makedirs(temp_dir, exist_ok=True)
        temp_path = os.path.join(temp_dir, "temp_upload.tmp")
        with open(temp_path, "wb") as f:
            f.write(contents)
            
        request.session["uploaded_filename"] = filename
        request.session["preview_headers"] = headers
        request.session["preview_rows"] = rows[1:21]
        request.session["temp_file_path"] = temp_path
        
        raise RedirectException(url=str(request.url_for("admin.preview_file")))
    except RedirectException:
        raise
    except Exception as e:
        import traceback
        traceback.print_exc()
        flash(request, f"Error processing file: {e}")
        raise RedirectException(url=str(request.url_for("admin.data_management")))


@router.get("/preview-file", name="admin.preview_file")
@admin_login_required
def preview_file(request: Request):
    filename = request.session.get("uploaded_filename")
    headers = request.session.get("preview_headers")
    preview_rows = request.session.get("preview_rows")
    
    if not filename or not headers or preview_rows is None:
        flash(request, "No preview data found. Please upload a file first.")
        raise RedirectException(url=str(request.url_for("admin.data_management")))
        
    db = get_db()
    cursor = db.execute("PRAGMA table_info(products)")
    columns = cursor.fetchall()
    db_fields = [row["name"] for row in columns if row["name"] != "id"]
    
    templates = request.app.state.templates
    return templates.TemplateResponse(
        request=request,
        name="admin/preview_file.html",
        context={
            "filename": filename,
            "headers": headers,
            "preview_rows": preview_rows,
            "db_fields": db_fields
        }
    )


@router.post("/import-data", name="admin.import_data")
@admin_login_required
async def import_data(request: Request):
    filename = request.session.get("uploaded_filename")
    temp_path = request.session.get("temp_file_path")
    
    if not filename or not temp_path or not os.path.exists(temp_path):
        flash(request, "Upload session expired or no file found. Please upload again.")
        raise RedirectException(url=str(request.url_for("admin.data_management")))
        
    form_data = await request.form()
    mappings = {}
    for k, v in form_data.items():
        if k.startswith("map_") and v:
            field_name = k[4:]
            mappings[field_name] = v
            
    required_fields = ["name", "category", "price", "stock", "aisle"]
    missing_required = [f for f in required_fields if f not in mappings]
    if missing_required:
        flash(request, f"Error: All core fields ({', '.join(required_fields)}) must be mapped. Missing: {', '.join(missing_required)}")
        raise RedirectException(url=str(request.url_for("admin.preview_file")))
        
    try:
        with open(temp_path, "rb") as f:
            contents = f.read()
            
        rows = []
        if filename.endswith(".xlsx"):
            wb = openpyxl.load_workbook(io.BytesIO(contents), data_only=True)
            sheet = wb.active
            for row in sheet.iter_rows(values_only=True):
                if any(cell is not None for cell in row):
                    rows.append([str(cell) if cell is not None else "" for cell in row])
        else:
            decoded = contents.decode("utf-8", errors="ignore")
            csv_reader = csv.reader(io.StringIO(decoded))
            rows = list(csv_reader)
            
        if not rows:
            flash(request, "The file contains no data.")
            raise RedirectException(url=str(request.url_for("admin.data_management")))
            
        headers = [h.strip() for h in rows[0] if h is not None]
        data_rows = rows[1:]
        
        db = get_db()
        cursor = db.execute("PRAGMA table_info(products)")
        columns_info = {row["name"]: row["type"] for row in cursor.fetchall()}
        
        imported = 0
        skipped = 0
        total_rows = len(data_rows)
        
        for row in data_rows:
            row_dict = {}
            for idx, h in enumerate(headers):
                if idx < len(row):
                    row_dict[h] = row[idx]
                else:
                    row_dict[h] = ""
                    
            db_values = {}
            valid = True
            
            for field, header in mappings.items():
                val_str = row_dict.get(header, "").strip()
                field_type = columns_info.get(field, "TEXT").upper()
                
                if field in required_fields and not val_str:
                    valid = False
                    break
                    
                if val_str == "":
                    db_values[field] = None
                    continue
                    
                try:
                    if field_type == "INTEGER":
                        db_values[field] = int(float(val_str))
                    elif field_type == "REAL":
                        db_values[field] = float(val_str)
                    else:
                        db_values[field] = val_str
                except ValueError:
                    valid = False
                    break
                    
            if valid:
                if db_values.get("price") is not None and db_values["price"] < 0:
                    valid = False
                if db_values.get("stock") is not None and db_values["stock"] < 0:
                    valid = False
                    
            if not valid:
                skipped += 1
                continue
                
            columns_str = ", ".join(db_values.keys())
            placeholders = ", ".join(["?"] * len(db_values))
            sql = f"INSERT INTO products ({columns_str}) VALUES ({placeholders})"
            db.execute(sql, list(db_values.values()))
            imported += 1
            
        db.execute(
            "INSERT INTO uploaded_files (filename, upload_date) VALUES (?, ?)",
            (filename, datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
        )
        db.commit()
        
        request.session["import_summary"] = {
            "filename": filename,
            "total": total_rows,
            "imported": imported,
            "skipped": skipped
        }
        
        try:
            if os.path.exists(temp_path):
                os.remove(temp_path)
        except Exception:
            pass
            
        request.session.pop("uploaded_filename", None)
        request.session.pop("preview_headers", None)
        request.session.pop("preview_rows", None)
        request.session.pop("temp_file_path", None)
        
        flash(request, "Import processing completed successfully.")
        raise RedirectException(url=str(request.url_for("admin.data_management")))
    except RedirectException:
        raise
    except Exception as e:
        import traceback
        traceback.print_exc()
        flash(request, f"Error during import: {e}")
        raise RedirectException(url=str(request.url_for("admin.data_management")))


@router.post("/add-column", name="admin.add_column")
@admin_login_required
async def add_column(request: Request):
    form_data = await request.form()
    column_name = form_data.get("column_name", "").strip().lower()
    data_type = form_data.get("data_type", "TEXT").strip()
    
    if not column_name:
        flash(request, "Column name is required.")
        raise RedirectException(url=str(request.url_for("admin.data_management")))
        
    if not re.match(r"^[a-z_][a-z0-9_]*$", column_name):
        flash(request, "Invalid column name format. Use lowercase letters, numbers, and underscores only. Must start with a letter.")
        raise RedirectException(url=str(request.url_for("admin.data_management")))
        
    db = get_db()
    cursor = db.execute("PRAGMA table_info(products)")
    existing = [row["name"].lower() for row in cursor.fetchall()]
    
    if column_name in existing:
        flash(request, f"Column '{column_name}' already exists in products table.")
        raise RedirectException(url=str(request.url_for("admin.data_management")))
        
    try:
        db.execute(f"ALTER TABLE products ADD COLUMN {column_name} {data_type}")
        db.execute("INSERT OR IGNORE INTO display_settings (field_name, is_visible) VALUES (?, 1)", (column_name,))
        db.execute("INSERT OR IGNORE INTO worker_field_permissions (field_name, can_view, can_edit) VALUES (?, 1, 0)", (column_name,))
        db.commit()
        
        flash(request, f"Custom column '{column_name}' ({data_type}) created successfully.")
    except Exception as e:
        flash(request, f"Error adding custom column: {e}")
        
    raise RedirectException(url=str(request.url_for("admin.data_management")))


@router.post("/display-settings", name="admin.display_settings")
@admin_login_required
async def display_settings_post(request: Request):
    form_data = await request.form()
    
    db = get_db()
    settings = query_db("SELECT field_name FROM display_settings")
    
    db.execute("UPDATE display_settings SET is_visible = 0")
    for row in settings:
        field = row["field_name"]
        if form_data.get(f"visible_{field}"):
            db.execute("UPDATE display_settings SET is_visible = 1 WHERE field_name = ?", (field,))
            
    db.commit()
    flash(request, "Customer visibility settings updated successfully.")
    raise RedirectException(url=str(request.url_for("admin.data_management")))


@router.api_route("/worker-permissions", methods=["GET", "POST"], name="admin.worker_permissions")
@admin_login_required
async def worker_permissions(request: Request):
    db = get_db()
    
    cursor = db.execute("PRAGMA table_info(products)")
    columns = [row["name"] for row in cursor.fetchall() if row["name"] != "id"]
    for col in columns:
        db.execute("INSERT OR IGNORE INTO display_settings (field_name, is_visible) VALUES (?, 1)", (col,))
        can_edit = 1 if col == "stock" else 0
        db.execute("INSERT OR IGNORE INTO worker_field_permissions (field_name, can_view, can_edit) VALUES (?, 1, ?)", (col, can_edit))
    db.commit()
    
    if request.method == "POST":
        form_data = await request.form()
        
        db.execute("UPDATE worker_field_permissions SET can_view = 0, can_edit = 0")
        db.execute("UPDATE worker_action_permissions SET is_enabled = 0")
        
        for k in form_data.keys():
            if k.startswith("view_"):
                field_name = k[5:]
                db.execute("UPDATE worker_field_permissions SET can_view = 1 WHERE field_name = ?", (field_name,))
            elif k.startswith("edit_"):
                field_name = k[5:]
                db.execute("UPDATE worker_field_permissions SET can_edit = 1 WHERE field_name = ?", (field_name,))
            elif k.startswith("action_"):
                action_name = k[7:]
                db.execute("UPDATE worker_action_permissions SET is_enabled = 1 WHERE action_name = ?", (action_name,))
                
        db.commit()
        flash(request, "Worker permissions updated successfully.")
        raise RedirectException(url=str(request.url_for("admin.worker_permissions")))
        
    field_perms = query_db("SELECT * FROM worker_field_permissions")
    action_perms = query_db("SELECT * FROM worker_action_permissions")
    
    templates = request.app.state.templates
    return templates.TemplateResponse(
        request=request,
        name="admin/worker_permissions.html",
        context={
            "field_perms": field_perms,
            "action_perms": action_perms
        }
    )
